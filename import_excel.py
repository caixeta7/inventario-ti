import sys
import uuid
import openpyxl
from datetime import datetime, date
from database import init_db, get_connection, get_all_sheet_names
from config import XLSX_PATH

HEADER_MAP = {
    "EQUIPAMENTO": "equipamento",
    "MARCA": "marca",
    "MODELO": "modelo",
    "MEMÓRIA": "memoria",
    "MEMORIA": "memoria",
    "HD": "hd",
    "POLEGADAS": "polegadas",
    "OWNER": "owner",
    "PATRIMONIO": "patrimonio",
    "SERVICE TAG": "service_tag",
    "HOSTNAME": "hostname",
    "LOCALIDADE": "localidade",
    "ANDAR": "andar",
    "SIGLA": "sigla",
    "USUÁRIO": "usuario",
    "USUARIO": "usuario",
    "CARGO": "cargo",
    "ÁREA": "area",
    "AREA": "area",
    "AQUISIÇÃO": "aquisicao",
    "AQUISICAO": "aquisicao",
    "VENCIMENTO DE GARANTIA": "vencimento_garantia",
    "OBSERVAÇÕES": "observacoes",
    "OBSERVACOES": "observacoes",
    "CONTRATO": "contrato",
    "INÍCIO DO CONTRATO": "inicio_contrato",
    "INICIO DO CONTRATO": "inicio_contrato",
    "VENCIMENTO DO CONTRATO": "vencimento_contrato",
    "LINHA ATUALIZADA EM": "linha_atualizada_em",
    "BAIXA EM": "baixa_em",
}

DATE_COLS = {"aquisicao", "vencimento_garantia", "inicio_contrato", "vencimento_contrato", "baixa_em"}

def parse_date(value):
    if not value:
        return None
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, str):
        val = value.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(val, fmt).strftime("%Y-%m-%d")
            except ValueError:
                pass
    return None

def import_planilha(path=None):
    file_path = path or XLSX_PATH
    print(f"Lendo planilha de: {file_path}")

    init_db()

    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Abas a importar: lê do banco (que já tem as 4 padrão na primeira inicialização)
    sheet_names = get_all_sheet_names()

    conn = get_connection()
    cursor = conn.cursor()

    # Limpa dados antigos
    cursor.execute("DELETE FROM ativos;")
    conn.commit()

    total_imported = 0

    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            print(f"  Aviso: Aba '{sheet_name}' não encontrada na planilha. Pulando...")
            continue

        ws = wb[sheet_name]
        header_row = [cell.value for cell in ws[1]]

        col_mapping = {}
        for idx, header in enumerate(header_row):
            if header:
                clean = str(header).strip().upper()
                if clean in HEADER_MAP and idx not in col_mapping:
                    col_mapping[idx] = HEADER_MAP[clean]

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        sheet_count = 0

        for row in rows:
            record = {
                "id": str(uuid.uuid4()),
                "sheet_name": sheet_name,
                "equipamento": None, "marca": None, "modelo": None, "memoria": None,
                "hd": None, "polegadas": None, "owner": None, "patrimonio": None,
                "service_tag": None, "hostname": None, "localidade": None, "andar": None,
                "sigla": None, "usuario": None, "cargo": None, "area": None,
                "aquisicao": None, "vencimento_garantia": None, "observacoes": None,
                "contrato": None, "inicio_contrato": None, "vencimento_contrato": None,
                "linha_atualizada_em": None, "baixa_em": None
            }

            has_data = False
            for idx, val in enumerate(row):
                if idx in col_mapping and val is not None:
                    col_name = col_mapping[idx]
                    if col_name in DATE_COLS:
                        parsed = parse_date(val)
                        if parsed:
                            record[col_name] = parsed
                            has_data = True
                    else:
                        s = str(val).strip()
                        if s:
                            record[col_name] = s
                            has_data = True

            if has_data:
                cursor.execute("""
                INSERT INTO ativos (
                    id, sheet_name, equipamento, marca, modelo, memoria, hd, polegadas, owner, patrimonio,
                    service_tag, hostname, localidade, andar, sigla, usuario, cargo, area,
                    aquisicao, vencimento_garantia, observacoes, contrato, inicio_contrato,
                    vencimento_contrato, linha_atualizada_em, baixa_em
                ) VALUES (
                    :id, :sheet_name, :equipamento, :marca, :modelo, :memoria, :hd, :polegadas, :owner, :patrimonio,
                    :service_tag, :hostname, :localidade, :andar, :sigla, :usuario, :cargo, :area,
                    :aquisicao, :vencimento_garantia, :observacoes, :contrato, :inicio_contrato,
                    :vencimento_contrato, :linha_atualizada_em, :baixa_em
                )
                """, record)
                sheet_count += 1
                total_imported += 1

        conn.commit()
        print(f"  -> Aba '{sheet_name}': {sheet_count} registros importados.")

    conn.close()
    print(f"\nSucesso! {total_imported} registros importados de {len(sheet_names)} abas.")

if __name__ == "__main__":
    filepath = sys.argv[1] if len(sys.argv) > 1 else None
    import_planilha(filepath)
