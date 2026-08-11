import os
import threading
import datetime
import openpyxl
from typing import List, Dict, Any
from config import XLSX_PATH
from database import get_connection, dict_from_rows, get_sheet_configs, get_all_sheet_names

_sync_lock = threading.Lock()
_is_processing = False
_pending_rerun = False

def enqueue_sync() -> None:
    conn = get_connection()
    conn.execute("INSERT INTO sync_queue DEFAULT VALUES;")
    conn.commit()
    conn.close()
    global _is_processing
    with _sync_lock:
        if not _is_processing:
            threading.Thread(target=_process_queue, daemon=True).start()
        else:
            global _pending_rerun
            _pending_rerun = True

def _process_queue() -> None:
    global _is_processing, _pending_rerun
    while True:
        with _sync_lock:
            if _is_processing:
                _pending_rerun = True
                return
            _is_processing = True
            _pending_rerun = False
        try:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT id FROM sync_queue WHERE processado = 0 ORDER BY id")
            pending = cursor.fetchall()
            if pending:
                pending_ids = [row["id"] for row in pending]
                cursor.execute("SELECT * FROM ativos WHERE deletado = 0 ORDER BY sheet_name, criado_em")
                ativos = dict_from_rows(cursor.fetchall())
                conn.close()
                _write_excel_all_sheets(ativos)
                conn = get_connection()
                placeholders = ",".join("?" * len(pending_ids))
                conn.execute(f"UPDATE sync_queue SET processado = 1, erro = NULL WHERE id IN ({placeholders})", pending_ids)
                conn.commit()
                conn.close()
            else:
                conn.close()
        except Exception as e:
            error_msg = str(e)
            print(f"[excel_sync] Erro ao sincronizar Excel: {error_msg}")
            try:
                conn = get_connection()
                conn.execute("UPDATE sync_queue SET erro = ? WHERE processado = 0", (error_msg,))
                conn.commit()
                conn.close()
            except Exception:
                pass
        finally:
            with _sync_lock:
                _is_processing = False
                if _pending_rerun:
                    _pending_rerun = False
                    continue
                break

def _write_excel_all_sheets(ativos: List[Dict[str, Any]]) -> None:
    target_path = XLSX_PATH
    os.makedirs(os.path.dirname(os.path.abspath(target_path)), exist_ok=True)

    if os.path.exists(target_path):
        try:
            wb = openpyxl.load_workbook(target_path)
        except Exception:
            wb = openpyxl.Workbook()
    else:
        wb = openpyxl.Workbook()

    #Abas gerenciadas dinamicamente — lidas do banco
    sheet_configs = get_sheet_configs()
    sheet_names = get_all_sheet_names()

    # Agrupa ativos por aba
    ativos_por_aba: Dict[str, List[Dict[str, Any]]] = {}
    for ativo in ativos:
        sheet = ativo.get("sheet_name", "Ativos")
        if sheet not in ativos_por_aba:
            ativos_por_aba[sheet] = []
        ativos_por_aba[sheet].append(ativo)

    # Reescreve cada aba gerenciada
    for sheet_name in sheet_names:
        config = sheet_configs.get(sheet_name)
        if not config:
            continue
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = wb.create_sheet(title=sheet_name)
        headers = [col[0] for col in config["columns"]]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=1, column=col_idx).font = openpyxl.styles.Font(bold=True)
        for ativo in ativos_por_aba.get(sheet_name, []):
            row_values = []
            for label, key in config["columns"]:
                val = ativo.get(key)
                if key in config["date_keys"] and val:
                    try:
                        val = datetime.datetime.strptime(str(val)[:10], "%Y-%m-%d").date()
                    except Exception:
                        pass
                row_values.append(val if val is not None else "")
            ws.append(row_values)

    # Escrita com fallback para permissão negada no Windows
    tmp_path = target_path + ".tmp"
    wb.save(tmp_path)
    wb.close()
    try:
        os.replace(tmp_path, target_path)
    except PermissionError:
        try:
            wb2 = openpyxl.load_workbook(tmp_path)
            wb2.save(target_path)
            wb2.close()
            os.remove(tmp_path)
        except Exception:
            pass

def get_sync_status() -> Dict[str, Any]:
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) as n FROM sync_queue WHERE processado = 0")
    pending_count = cursor.fetchone()["n"]
    cursor.execute("SELECT erro FROM sync_queue WHERE erro IS NOT NULL ORDER BY id DESC LIMIT 1")
    last_err_row = cursor.fetchone()
    last_error = last_err_row["erro"] if last_err_row else None
    conn.close()
    file_ok = os.path.exists(XLSX_PATH) or os.access(os.path.dirname(os.path.abspath(XLSX_PATH)), os.W_OK)
    return {
        "pendingCount": pending_count,
        "lastError": last_error,
        "adapter": "Excel local (openpyxl)",
        "destinoOk": file_ok,
        "destinoDetalhe": "Planilha sincronizada automaticamente" if file_ok else "Aguardando criação do arquivo"
    }
